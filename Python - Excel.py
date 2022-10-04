from openpyxl import *

# Lendo a planilha
planilha = load_workbook(
    filename="E:\Programas em python\Python com Excel\TimeSheet-ThacioLopes-2022-08.xlsx")

# Mexendo na aba Time-Sheet
timeSheet = planilha['Time-Sheet']


#Entrada de dados
# mes = input('Time Sheet referente de qual mês?: ')


# Mudar o mes
alterar_o_mes = input('Você gostaria de alterar o mês?: ')
if alterar_o_mes == 'sim':
    print('Mudando O mes')
    mes = input('Time Sheet referente de qual mês?: ')
    alterando_o_mes = timeSheet.cell(row=2, column=2).value = mes

elif alterar_o_mes == 'nao':
    print('Não ira alterar o mês')


# semana = int(input('Qual semana estamos? Ex: 1 , 2, 3, 4 : '))


def inserir_dia():
    dia = int(input('Qual e o dia: ')) - 1
    mes = int(input('Qual e o mes: '))
    ano = 22
    resultado = ''

    for item in range(5):
        dia = dia + 1
        if dia <= 3:
            dia - 1

        # print('Tipo do dia: ',type(dia))

        result = f'{dia}/{mes}/{ano} {dia}/{mes}/{ano} '

        resultado = resultado + result

    valorData = resultado.split()



    print(valorData)



    """
    for index in range(len(valorData)):
        #print(index,valorData[index])

        print(teste)
    """

    def inserindoNoExcel(valor):
        semana = int(input('Qual semana estamos? Ex: 1 , 2, 3, 4 : '))

        if semana == 1:

        elif semana == 2:
            ##Colocar a row de 10 a A 20 Pulando 10 linhas
            print('Semana 2')

        elif semana == 3:
            ##Colocar a row de 20 a A 30 Pulando 10 linhas
            print('Semana 3')

        elif semana == 3:
            ##Colocar a row de 30 a A 40 Pulando 10 linhas
            print('Semana 3')

        elif semana == 4:
            ##Colocar a row de 40 a A 50 Pulando 10 linhas
            print('Semana 4')

        else:
            print('Semana invalida')

    inserindoNoExcel(valorData[9])


alterando_as_ceululas = timeSheet.cell(
    row=5, column=1).value = str(inserir_dia())

print(f'NA PLANILHA: ,{alterando_as_ceululas.value}')


# Se a semana for 1 não pular

# se a semena for 2 pular 10 linhas

# Se a semana for 3 não pular 20 linhas

# se a semena for 4 pular 30 linhas

# fazer as a data por 5x OK












# 01/08/22
primeiro_dia_da_semana = input('que dia foi sua segunda feira?: ')
fim= primeiro_dia_da_semana + 8


while (dias<=fim):
    print(dias)
    dias= dias + 1

ano = 2022

print(f'{primeiro_dia_da_semana}/ {mes} /{ano}')

# print(timeSheet.cell(row=2, column=2 ).value)

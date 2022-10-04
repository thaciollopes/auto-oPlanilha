from openpyxl import *

dias= 1
fim= dias + 3
mes= 8
ano = 22
# print(fim)

planilha = load_workbook(
    filename="E:\Programas em python\Python com Excel\TimeSheet-ThacioLopes-2022-08.xlsx")

# Mexendo na aba Time-Sheet
timeSheet = planilha['Time-Sheet']

datas= [ '1/1/22', '1/1/22', '2/1/22', '2/1/22', '3/1/22', '3/1/22', '4/1/22', '4/1/22', '5/1/22', '5/1/22']

var = 1
print(datas[var])

linha = 0

#Adiciondo nas linhas
while linha <= 10:
    teste = (datas[linha])
    alterando = timeSheet.cell(row=5 + linha, column=1, value=teste)
    print('---'*10)
    linha = linha + 1

    if linha == 10:
        break


# alterando = timeSheet.cell(row=linha, column=2, value=datas)

t = 4

# d = timeSheet['A4'] = 'TESTEEEE'



# print(f'teste :  {vl}')

c = timeSheet[f'A{t}']
print(f'valor é {c.value}')





planilha.save('teste.xlsx')



# alterando_o_mes = timeSheet.cell(row=linha, column=2).value = datas[linha]

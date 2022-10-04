from openpyxl import *
import xlwings as xw

mes = input('Time Sheet referente de qual mês?: ')

planilhaJaCriada = input('A planilha já foi criada ? Ex: s, n : ')

if planilhaJaCriada == 's':
    print(f'E:\Programas em python\Python com Excel\Planilha_{mes}.xlsx')
    planilha = load_workbook(filename= f'E:\Programas em python\Python com Excel\Planilha_{mes}.xlsx',read_only=False, keep_vba=True)



elif planilhaJaCriada == 'NAO':
      planilha = load_workbook(filename="E:\Programas em python\Python com Excel\TimeSheet-ThacioLopes-2022-BASE.xlsx")


# Lendo a planilha NOVA
planilha = load_workbook(filename="E:\Programas em python\Python com Excel\TimeSheet-ThacioLopes-2022-BASE.xlsx")

# Mexendo na aba Time-Sheet
timeSheet = planilha['Time-Sheet']



alterando_o_mes_de_referencia = timeSheet.cell(row=2, column=2).value = mes


#Var global
linha = 0
#Var global
posicaoDaSemana = 0

semana = int(input('Qual semana estamos? Ex: 1 , 2, 3, 4 : '))

if semana == 1:
    posicaoDaSemana = 5


elif semana == 2:
    ##Colocar a row de 10 a A 20 Pulando 10 linhas
    print('Semana 2')
    posicaoDaSemana = 15

elif semana == 3:
    ##Colocar a row de 20 a A 30 Pulando 10 linhas
    print('Semana 3')
    posicaoDaSemana = 25

elif semana == 3:
    ##Colocar a row de 30 a A 40 Pulando 10 linhas
    print('Semana 3')
    posicaoDaSemana = 35

elif semana == 4:
    ##Colocar a row de 40 a A 50 Pulando 10 linhas
    print('Semana 4')
    posicaoDaSemana = 45

else:
    print('Semana invalida')


def inserir_dia():
    dia = int(input('Qual e o dia: ')) - 1
    mes = int(input('Qual e o mes: '))
    ano = 22
    resultado = ''
    global linha
    global posicaoDaSemana

    for item in range(5):
        dia = dia + 1
        if dia <= 3:
            dia - 1


        result = f'{dia}/{mes}/{ano} {dia}/{mes}/{ano} '

        resultado = resultado + result

    valorData = resultado.split()



    print(valorData)
    print('------'*10)

    while linha <= 10:
        valorDaData = (valorData[linha])
        add_Planilha = timeSheet.cell(row=posicaoDaSemana + linha, column=1, value=valorDaData)
        print('---' * 10)
        linha = linha + 1

        if linha == 10:
            break



inserir_dia()


#Salvando A Planilha

planilha.save(f'Planilha_{mes}.xlsx')
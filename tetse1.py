# Abrindo a planilha
from openpyxl import *

# lendo a planilha
planilha = load_workbook(
    filename="E:\Programas em python\Python com Excel\TimeSheet-ThacioLopes-2022-08.xlsx")

# Mexendo na aba Time-Sheet
timeSheet = planilha['Time-Sheet']


def inserir_dia():
    dia = int(input('Qual e o dia: ')) - 1
    mes = int(input('Qual e o mes: '))
    ano = 22
    resultado = ''

    for item in range(5):
        dia = dia + 1
        if dia <= 3:
            dia - 1

        # print('Tipo do dia: ',type(dia))

        result = f'{dia}/{mes}/{ano} {dia}/{mes}/{ano} '

        resultado = resultado + result

    valorData = resultado.split()

    print(valorData)

    """
    for index in range(len(valorData)):
        #print(index,valorData[index])
        
        print(teste)
    """
    def inserindoNoExcel(valor):
        semana = int(input('Qual semana estamos? Ex: 1 , 2, 3, 4 : '))
        
        if semana == 1:
            ##Colocar a row de 0 a A 10 Pulando 10 linhas
            print('Semana 1')
            row = 0
            for row in range(10):
                index = 0
                '''
                timeSheet.cell(row=5, column=1).value = valor[index]
                print(valor)
                '''
                print(valor[0])
                row = row + 1
              
                print(f'valor da index: {index}')
                print(f'valor da row: {row}')
            
        elif semana == 2:
            ##Colocar a row de 10 a A 20 Pulando 10 linhas
            print('Semana 2')

        elif semana == 3:
            ##Colocar a row de 20 a A 30 Pulando 10 linhas
            print('Semana 3')

        elif semana == 3:
            ##Colocar a row de 30 a A 40 Pulando 10 linhas
            print('Semana 3')

        elif semana == 4:
            ##Colocar a row de 40 a A 50 Pulando 10 linhas
            print('Semana 4')

        else:
            print('Semana invalida')
    
    inserindoNoExcel(valorData[9])


alterando_as_ceululas = timeSheet.cell(
    row=5, column=1).value = str(inserir_dia())
# print('NA PLANILHA: ',alterando_as_ceululas)


'''
se o valor da planilha for A1 ele adiciona na planilha 


'''

# teste = timeSheet.cell(row=1, column=2).value = 'TTTT'
# print(teste)
